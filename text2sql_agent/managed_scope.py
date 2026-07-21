"""Request-scoped managed-company list parsing for Athena queries.

The uploaded values are intentionally kept out of database staging tables.  A
validated list is carried with the current request and rendered as a bounded
Athena ``VALUES`` relation by :mod:`text2sql_agent.tools.sql_builders`.
"""

from __future__ import annotations

import csv
import io
import math
import re
import zipfile
from pathlib import Path
from typing import Any, Iterable


MAX_MANAGED_COMPANIES = 5_000
MAX_MANAGED_SCOPE_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_MANAGED_SCOPE_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MANAGED_SCOPE_PARAMETER = "관리기업목록"
SUPPORTED_MANAGED_SCOPE_EXTENSIONS = {".txt", ".csv", ".xlsx", ".xlsm"}

_BUSINESS_NUMBER_PATTERN = re.compile(
    r"(?<!\d)(?:\d{3}\s*-\s*\d{2}\s*-\s*\d{5}|\d{10})(?!\d)"
)
_SQL_META_PATTERN = re.compile(r"(?:;|--|/\*|\*/|['\"`()])")
_HEADER_ALIASES = {
    "사업자등록번호",
    "사업자번호",
    "사업자등록번호목록",
    "businessregistrationnumber",
    "businessnumber",
    "bizno",
}


class ManagedScopeParseError(ValueError):
    """Raised when a managed-company input cannot be validated safely."""


def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s_\-().]", "", str(value or "")).lower()


def _is_business_number_header(value: Any) -> bool:
    return _normalized_header(value) in _HEADER_ALIASES


def normalize_business_number(value: Any) -> str:
    """Return a ten-digit business registration number.

    Identifiers may be supplied as Excel numbers or as text with ordinary
    hyphen/space formatting.  No checksum rule is imposed because the source
    systems may contain historical identifiers; this boundary validates format
    and prevents arbitrary SQL fragments.
    """

    if value is None or isinstance(value, bool):
        raise ManagedScopeParseError("사업자등록번호가 비어 있습니다.")
    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise ManagedScopeParseError("사업자등록번호는 소수점 없는 10자리 값이어야 합니다.")
        raw = str(int(value))
    else:
        raw = str(value).strip()
        if raw.startswith("'") and not raw.endswith("'"):
            # Excel users sometimes prefix a value with an apostrophe to force
            # text.  openpyxl normally removes it, but TXT exports may retain it.
            raw = raw[1:]
    compact = re.sub(r"[\s-]", "", raw)
    if not re.fullmatch(r"\d{10}", compact):
        raise ManagedScopeParseError("사업자등록번호는 하이픈 유무와 관계없이 숫자 10자리여야 합니다.")
    return compact


def _deduplicate_and_bound(values: Iterable[Any], *, max_items: int) -> tuple[list[str], int]:
    numbers: list[str] = []
    seen: set[str] = set()
    duplicate_count = 0
    for value in values:
        number = normalize_business_number(value)
        if number in seen:
            duplicate_count += 1
            continue
        seen.add(number)
        numbers.append(number)
        if len(numbers) > max_items:
            raise ManagedScopeParseError(f"관리기업은 한 번에 최대 {max_items:,}개까지 조회할 수 있습니다.")
    if not numbers:
        raise ManagedScopeParseError("유효한 사업자등록번호가 없습니다.")
    return numbers, duplicate_count


def parse_business_number_list(value: Any, *, max_items: int = MAX_MANAGED_COMPANIES) -> list[str]:
    """Parse pasted text or an iterable into a normalized, unique list."""

    if isinstance(value, (list, tuple, set)):
        numbers, _ = _deduplicate_and_bound(value, max_items=max_items)
        return numbers
    if value is None:
        raise ManagedScopeParseError("관리기업 목록이 비어 있습니다.")

    text = str(value).strip()
    if not text:
        raise ManagedScopeParseError("관리기업 목록이 비어 있습니다.")
    if _SQL_META_PATTERN.search(text):
        raise ManagedScopeParseError("관리기업 목록에는 사업자등록번호와 쉼표·줄바꿈만 입력해 주세요.")

    matches = _BUSINESS_NUMBER_PATTERN.findall(text)
    if not matches:
        raise ManagedScopeParseError("10자리 사업자등록번호를 찾지 못했습니다.")
    numbers, _ = _deduplicate_and_bound(matches, max_items=max_items)
    return numbers


def render_athena_business_number_values(value: Any) -> str:
    """Render a digits-only list as Athena/Trino row values."""

    numbers = parse_business_number_list(value)
    return ", ".join(f"('{number}')" for number in numbers)


def _decode_text_upload(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ManagedScopeParseError("TXT/CSV 파일은 UTF-8 또는 CP949 인코딩이어야 합니다.")


def _parse_csv_upload(content: bytes) -> tuple[list[str], int, str]:
    text = _decode_text_upload(content)
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise ManagedScopeParseError("업로드한 파일이 비어 있습니다.")

    header_row = -1
    target_col = -1
    for row_index, row in enumerate(rows[:20]):
        for col_index, value in enumerate(row):
            if _is_business_number_header(value):
                header_row, target_col = row_index, col_index
                break
        if target_col >= 0:
            break
    if target_col < 0:
        # Headerless CSV/TXT exports are accepted using the first column.
        header_row, target_col = -1, 0

    raw_values: list[Any] = []
    invalid_rows: list[int] = []
    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        value = row[target_col] if target_col < len(row) else None
        if value is None or not str(value).strip() or str(value).lstrip().startswith("#"):
            continue
        try:
            raw_values.append(normalize_business_number(value))
            if len(raw_values) > MAX_MANAGED_COMPANIES:
                raise ManagedScopeParseError(
                    f"관리기업은 한 번에 최대 {MAX_MANAGED_COMPANIES:,}행까지 업로드할 수 있습니다."
                )
        except ManagedScopeParseError:
            if len(raw_values) > MAX_MANAGED_COMPANIES:
                raise
            invalid_rows.append(row_number)
    if invalid_rows:
        shown = ", ".join(map(str, invalid_rows[:8]))
        suffix = " 외" if len(invalid_rows) > 8 else ""
        raise ManagedScopeParseError(f"사업자등록번호 형식이 잘못된 행이 있습니다: {shown}{suffix}")
    numbers, duplicates = _deduplicate_and_bound(raw_values, max_items=MAX_MANAGED_COMPANIES)
    return numbers, duplicates, "텍스트"


def _parse_txt_upload(content: bytes) -> tuple[list[str], int, str]:
    text = _decode_text_upload(content)
    lines = [line.strip() for line in text.splitlines() if line.strip() and not line.lstrip().startswith("#")]
    if lines and _is_business_number_header(lines[0].split(",", 1)[0]):
        lines = lines[1:]
    if not lines:
        raise ManagedScopeParseError("업로드한 파일이 비어 있습니다.")
    body = "\n".join(lines)
    if _SQL_META_PATTERN.search(body):
        raise ManagedScopeParseError("파일에는 사업자등록번호와 쉼표·줄바꿈만 입력해 주세요.")
    matches: list[str] = []
    invalid_rows: list[int] = []
    for row_number, line in enumerate(lines, start=1):
        row_matches = _BUSINESS_NUMBER_PATTERN.findall(line)
        remainder = _BUSINESS_NUMBER_PATTERN.sub("", line)
        remainder = re.sub(r"[\s,;|\t]", "", remainder)
        if not row_matches or remainder:
            invalid_rows.append(row_number)
            continue
        matches.extend(row_matches)
        if len(matches) > MAX_MANAGED_COMPANIES:
            raise ManagedScopeParseError(
                f"관리기업은 한 번에 최대 {MAX_MANAGED_COMPANIES:,}행까지 업로드할 수 있습니다."
            )
    if invalid_rows:
        shown = ", ".join(map(str, invalid_rows[:8]))
        suffix = " 외" if len(invalid_rows) > 8 else ""
        raise ManagedScopeParseError(f"사업자등록번호 형식이 잘못된 행이 있습니다: {shown}{suffix}")
    numbers, duplicates = _deduplicate_and_bound(matches, max_items=MAX_MANAGED_COMPANIES)
    return numbers, duplicates, "텍스트"


def _parse_excel_upload(content: bytes) -> tuple[list[str], int, str]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            if sum(info.file_size for info in archive.infolist()) > MAX_MANAGED_SCOPE_XLSX_UNCOMPRESSED_BYTES:
                raise ManagedScopeParseError("압축 해제된 Excel 파일 크기는 50MB 이하여야 합니다.")
    except ManagedScopeParseError:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise ManagedScopeParseError("Excel 파일 형식이 올바르지 않습니다.") from exc

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise ManagedScopeParseError("Excel 파서가 설치되어 있지 않습니다.") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ManagedScopeParseError("Excel 파일을 읽을 수 없습니다. 손상 여부와 파일 형식을 확인해 주세요.") from exc
    try:
        selected = None
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
                for col_index, value in enumerate(row, start=1):
                    if _is_business_number_header(value):
                        selected = (sheet, row_index, col_index)
                        break
                if selected:
                    break
            if selected:
                break
        if selected is None:
            raise ManagedScopeParseError(
                "Excel에서 '사업자등록번호' 열을 찾지 못했습니다. 제공된 예시 파일 형식을 사용해 주세요."
            )

        sheet, header_row, target_col = selected
        raw_values: list[Any] = []
        invalid_rows: list[int] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, min_col=target_col, max_col=target_col, values_only=True),
            start=header_row + 1,
        ):
            value = row[0] if row else None
            if value is None or not str(value).strip():
                continue
            try:
                raw_values.append(normalize_business_number(value))
                if len(raw_values) > MAX_MANAGED_COMPANIES:
                    raise ManagedScopeParseError(
                        f"관리기업은 한 번에 최대 {MAX_MANAGED_COMPANIES:,}행까지 업로드할 수 있습니다."
                    )
            except ManagedScopeParseError:
                if len(raw_values) > MAX_MANAGED_COMPANIES:
                    raise
                invalid_rows.append(row_number)
        if invalid_rows:
            shown = ", ".join(map(str, invalid_rows[:8]))
            suffix = " 외" if len(invalid_rows) > 8 else ""
            raise ManagedScopeParseError(f"'{sheet.title}' 시트의 사업자등록번호 형식 오류 행: {shown}{suffix}")
        numbers, duplicates = _deduplicate_and_bound(raw_values, max_items=MAX_MANAGED_COMPANIES)
        return numbers, duplicates, sheet.title
    finally:
        workbook.close()


def parse_managed_scope_upload(filename: str, content: bytes) -> dict[str, Any]:
    """Parse an uploaded TXT/CSV/XLSX file without persisting its contents."""

    safe_name = Path(filename or "").name
    extension = Path(safe_name).suffix.lower()
    if extension not in SUPPORTED_MANAGED_SCOPE_EXTENSIONS:
        if extension == ".xls":
            raise ManagedScopeParseError("구형 .xls 파일은 .xlsx로 저장한 뒤 업로드해 주세요.")
        allowed = ", ".join(sorted(SUPPORTED_MANAGED_SCOPE_EXTENSIONS))
        raise ManagedScopeParseError(f"지원하는 파일 형식은 {allowed} 입니다.")
    if not content:
        raise ManagedScopeParseError("업로드한 파일이 비어 있습니다.")
    if len(content) > MAX_MANAGED_SCOPE_UPLOAD_BYTES:
        raise ManagedScopeParseError("파일 크기는 5MB 이하여야 합니다.")

    if extension in {".xlsx", ".xlsm"}:
        numbers, duplicate_count, source_sheet = _parse_excel_upload(content)
    elif extension == ".csv":
        numbers, duplicate_count, source_sheet = _parse_csv_upload(content)
    else:
        numbers, duplicate_count, source_sheet = _parse_txt_upload(content)
    return {
        "parameter_name": MANAGED_SCOPE_PARAMETER,
        "business_numbers": numbers,
        "count": len(numbers),
        "duplicates_removed": duplicate_count,
        "source_filename": safe_name,
        "source_sheet": source_sheet,
    }
