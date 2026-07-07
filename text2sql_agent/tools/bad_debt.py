"""Bad debt cost-rate analysis tool and Excel workbook generation."""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import BAD_DEBT_OUTPUT_DIR, DB_SCHEMA, DB_SCHEMA_PREFIX
from ..db import execute_sql
from ..llm import _call_llm, _normalize_llm_text
from .sql_builders import _athena_partition_conds, _calc_months_back, _escape_like, _sanitize_param

# ---------------------------------------------------------------------------
# 5. 대손비용률 Tool 구현
# ---------------------------------------------------------------------------


BAD_DEBT_QUERIES = {
    "월초충당금": """WITH cor_table AS (
    SELECT DISTINCT "기업고객식별자"
    FROM card_system.tmdaa5d01
    WHERE LOWER("가맹점명") LIKE LOWER('%{가맹점명}%'){partition_tmdaa5d01_기준년월}
),
cm AS (
    SELECT 기준년월, 고객식별자,
        SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 원화대출잔액
    FROM card_system.tbmewcm94
    WHERE 기준년월 = '{월초년월}' AND 충당금차주구분코드 = '2'{partition_tbmewcm94_월초년월}
    GROUP BY 기준년월, 고객식별자
),
join_s AS (
    SELECT A.*,
        CASE WHEN B."기업고객식별자" IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM cm A LEFT JOIN cor_table B ON A.고객식별자 = B."기업고객식별자"
)
SELECT 기준년월, 구분,
    SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 잔액,
    ROUND(SUM(기대대손충당금) / NULLIF(SUM(원화대출잔액), 0) * 100, 2) AS 충당금적립률
FROM join_s GROUP BY 기준년월, 구분 ORDER BY 기준년월, 구분""",

    "월말충당금": """WITH cor_table AS (
    SELECT DISTINCT "기업고객식별자"
    FROM card_system.tmdaa5d01
    WHERE LOWER("가맹점명") LIKE LOWER('%{가맹점명}%'){partition_tmdaa5d01_기준년월}
),
cm AS (
    SELECT 기준년월, 고객식별자,
        SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 원화대출잔액
    FROM card_system.tbmewcm94
    WHERE 기준년월 = '{기준년월}' AND 충당금차주구분코드 = '2'{partition_tbmewcm94_기준년월}
    GROUP BY 기준년월, 고객식별자
),
join_s AS (
    SELECT A.*,
        CASE WHEN B."기업고객식별자" IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM cm A LEFT JOIN cor_table B ON A.고객식별자 = B."기업고객식별자"
)
SELECT 기준년월, 구분,
    SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 잔액,
    ROUND(SUM(기대대손충당금) / NULLIF(SUM(원화대출잔액), 0) * 100, 2) AS 충당금적립률
FROM join_s GROUP BY 기준년월, 구분 ORDER BY 기준년월, 구분""",

    "상각내역": """WITH cor_table AS (
    SELECT DISTINCT a.기업고객식별자
    FROM card_system.tbdaadt01 a
    WHERE a.가맹점상태구분코드 IN ('1', '01') AND LOWER(a.가맹점명) LIKE LOWER('%{가맹점명}%'){partition_tbdaadt01_기준년월}
),
sd06 AS (
    SELECT 고객식별자,
        SUM(특수채권편입원금) AS 특수채권편입원금, SUM(특수채권편입가지급금) AS 특수채권편입가지급금,
        SUM(특수채권편입원금 + 특수채권편입가지급금) AS 상각금액
    FROM card_system.tbmaisd06
    WHERE CAST(SUBSTR(특수채권편입기준년월일, 1, 6) AS INTEGER) BETWEEN CAST('{시작년월}' AS INTEGER) AND CAST('{기준년월}' AS INTEGER)
      AND 회계계정과목코드 = 'A040101010000' AND 개인기업구분코드 = '2'{partition_tbmaisd06_기간}
    GROUP BY 고객식별자
),
sd06_a AS (
    SELECT A.*,
        CASE WHEN B.기업고객식별자 IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM sd06 A LEFT JOIN cor_table B ON A.고객식별자 = B.기업고객식별자
)
SELECT 구분,
    SUM(특수채권편입원금) AS 특수채권편입원금, SUM(특수채권편입가지급금) AS 특수채권편입가지급금,
    SUM(상각금액) AS 상각금액
FROM sd06_a GROUP BY 구분 ORDER BY 구분""",

    "대손비용률종합": """WITH cor_table AS (
    SELECT DISTINCT "기업고객식별자"
    FROM card_system.tmdaa5d01
    WHERE LOWER("가맹점명") LIKE LOWER('%{가맹점명}%'){partition_tmdaa5d01_기준년월}
),
cm_start AS (
    SELECT 고객식별자, SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 원화대출잔액
    FROM card_system.tbmewcm94
    WHERE 기준년월 = '{월초년월}' AND 충당금차주구분코드 = '2'{partition_tbmewcm94_월초년월}
    GROUP BY 고객식별자
),
cm_start_flagged AS (
    SELECT A.고객식별자, A.기대대손충당금, A.원화대출잔액,
        CASE WHEN B."기업고객식별자" IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM cm_start A LEFT JOIN cor_table B ON A.고객식별자 = B."기업고객식별자"
),
cm_end AS (
    SELECT 고객식별자, SUM(기대대손충당금) AS 기대대손충당금, SUM(원화대출잔액) AS 원화대출잔액
    FROM card_system.tbmewcm94
    WHERE 기준년월 = '{기준년월}' AND 충당금차주구분코드 = '2'{partition_tbmewcm94_기준년월}
    GROUP BY 고객식별자
),
cm_end_flagged AS (
    SELECT A.고객식별자, A.기대대손충당금, A.원화대출잔액,
        CASE WHEN B."기업고객식별자" IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM cm_end A LEFT JOIN cor_table B ON A.고객식별자 = B."기업고객식별자"
),
cor_table_full AS (
    SELECT DISTINCT a.기업고객식별자 FROM card_system.tbdaadt01 a
    WHERE a.가맹점상태구분코드 IN ('1', '01') AND LOWER(a.가맹점명) LIKE LOWER('%{가맹점명}%'){partition_tbdaadt01_기준년월}
),
sd06 AS (
    SELECT 고객식별자, SUM(특수채권편입원금 + 특수채권편입가지급금) AS 상각금액
    FROM card_system.tbmaisd06
    WHERE CAST(SUBSTR(특수채권편입기준년월일, 1, 6) AS INTEGER) BETWEEN CAST('{시작년월}' AS INTEGER) AND CAST('{기준년월}' AS INTEGER)
      AND 회계계정과목코드 = 'A040101010000' AND 개인기업구분코드 = '2'{partition_tbmaisd06_기간}
    GROUP BY 고객식별자
),
sd06_flagged AS (
    SELECT A.고객식별자, A.상각금액,
        CASE WHEN B.기업고객식별자 IS NOT NULL THEN '1' ELSE '0' END AS 구분
    FROM sd06 A LEFT JOIN cor_table_full B ON A.고객식별자 = B.기업고객식별자
),
by_group AS (
    SELECT COALESCE(s.구분, e.구분, w.구분) AS 구분,
        COALESCE(SUM(s.기대대손충당금), 0) AS 월초_충당금, COALESCE(SUM(s.원화대출잔액), 0) AS 월초_잔액,
        COALESCE(SUM(e.기대대손충당금), 0) AS 월말_충당금, COALESCE(SUM(e.원화대출잔액), 0) AS 월말_잔액,
        COALESCE(SUM(w.상각금액), 0) AS 상각금액
    FROM cm_start_flagged s
    FULL OUTER JOIN cm_end_flagged e ON s.고객식별자 = e.고객식별자 AND s.구분 = e.구분
    FULL OUTER JOIN sd06_flagged w ON COALESCE(s.고객식별자, e.고객식별자) = w.고객식별자
        AND COALESCE(s.구분, e.구분) = w.구분
    GROUP BY COALESCE(s.구분, e.구분, w.구분)
),
with_cost AS (
    SELECT 구분, 월초_충당금, 월초_잔액, 월말_충당금, 월말_잔액, 상각금액,
        (월말_충당금 - 월초_충당금 + 상각금액) AS 대손비용,
        CASE WHEN (월초_잔액 + 월말_잔액) > 0
            THEN ROUND((월말_충당금 - 월초_충당금 + 상각금액) * 12.0 / ((월초_잔액 + 월말_잔액) / 2.0) * 100, 4)
            ELSE 0 END AS 대손비용률_퍼센트
    FROM by_group
),
total AS (
    SELECT SUM(대손비용) AS 전체_대손비용,
        SUM(월초_잔액) AS 전체_월초_잔액, SUM(월말_잔액) AS 전체_월말_잔액,
        CASE WHEN (SUM(월초_잔액) + SUM(월말_잔액)) > 0
            THEN ROUND(SUM(대손비용) * 12.0 / ((SUM(월초_잔액) + SUM(월말_잔액)) / 2.0) * 100, 4)
            ELSE 0 END AS 전체_대손비용률_퍼센트
    FROM with_cost
)
SELECT '{기준년월}' AS 기준년월, c.구분,
    c.월초_충당금, c.월초_잔액, c.월말_충당금, c.월말_잔액, c.상각금액,
    c.대손비용, c.대손비용률_퍼센트,
    CASE WHEN c.구분 = '1' THEN t.전체_대손비용률_퍼센트 ELSE NULL END AS 전체_대손비용률_퍼센트,
    CASE WHEN c.구분 = '1' AND t.전체_대손비용률_퍼센트 > 0
        THEN ROUND(c.대손비용률_퍼센트 / t.전체_대손비용률_퍼센트, 4) ELSE NULL END AS 보정계수
FROM with_cost c CROSS JOIN total t ORDER BY c.구분""",
}


def _to_float(value) -> float:
    """SQL 결과 셀을 float으로 변환한다. NULL/빈 값/변환 실패는 0.0으로 처리한다."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _partition_and(table_name: str, start: str, end: str = "", alias: str = "") -> str:
    try:
        conds = _athena_partition_conds(table_name, start=start, end=end or start, alias=alias)
    except Exception:
        return ""
    if not conds:
        return ""
    return "".join(f"\n      AND {cond}" for cond in conds)


def _build_bad_debt_sql(query_template: str, params: dict) -> str:
    merchant = _escape_like(params.get("가맹점명", ""))
    yyyymm = _sanitize_param(params.get("기준년월", ""))
    prev_month = _calc_months_back(yyyymm, 1) if yyyymm else ""
    start_month = _calc_months_back(yyyymm, 0) if yyyymm else ""
    sql = query_template.replace("{가맹점명}", merchant)
    sql = sql.replace("{기준년월}", yyyymm)
    sql = sql.replace("{월초년월}", prev_month)
    sql = sql.replace("{시작년월}", start_month)
    partition_replacements = {
        "{partition_tmdaa5d01_기준년월}": _partition_and("tmdaa5d01", yyyymm),
        "{partition_tbmewcm94_월초년월}": _partition_and("tbmewcm94", prev_month),
        "{partition_tbmewcm94_기준년월}": _partition_and("tbmewcm94", yyyymm),
        "{partition_tbdaadt01_기준년월}": _partition_and("tbdaadt01", yyyymm, alias="a"),
        "{partition_tbmaisd06_기간}": _partition_and("tbmaisd06", start_month, yyyymm),
    }
    for placeholder, condition in partition_replacements.items():
        sql = sql.replace(placeholder, condition)
    # 템플릿의 고정 prefix(card_system.)를 설정된 스키마 prefix로 치환한다.
    # 기본 설정(card_system)이면 변화 없음, 다른 스키마면 교체, 빈 스키마면 prefix 제거.
    if DB_SCHEMA != "card_system":
        sql = sql.replace("card_system.", DB_SCHEMA_PREFIX)
    return sql


def _generate_bad_debt_excel(
    params: dict,
    results: dict[str, tuple[list[str], list[tuple]]],
) -> str:
    os.makedirs(BAD_DEBT_OUTPUT_DIR, exist_ok=True)
    merchant = params.get("가맹점명", "기업")
    yyyymm = params.get("기준년월", "")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"대손비용률분석_{merchant}_{yyyymm}_{ts}.xlsx"
    filepath = os.path.join(BAD_DEBT_OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="맑은 고딕", size=10)
    num_font = Font(name="맑은 고딕", size=10)
    title_font = Font(name="맑은 고딕", bold=True, size=14, color="1F3864")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sheet_order = ["월초충당금", "월말충당금", "상각내역", "대손비용률종합"]
    for idx, sheet_name in enumerate(sheet_order):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        columns, rows = results.get(sheet_name, ([], []))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{merchant} 대손비용률 분석 - {sheet_name} ({yyyymm})"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        if not columns:
            ws.cell(row=3, column=1, value="조회 결과가 없습니다.").font = data_font
            continue
        header_row = 3
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)):
                    cell.font = num_font
                    cell.alignment = Alignment(horizontal="right")
                    col_name_lower = columns[col_idx - 1].lower()
                    if "보정계수" in col_name_lower:
                        cell.number_format = '#,##0.0000'
                    elif "퍼센트" in col_name_lower or "률" in col_name_lower or "율" in col_name_lower:
                        cell.number_format = '#,##0.0000"%"'
                    elif any(k in col_name_lower for k in ["금액", "한도", "이용", "연체", "잔액", "충당금", "상각", "원금", "가지급", "대손비용"]):
                        cell.number_format = '#,##0'
                else:
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left")
        for col_idx in range(1, len(columns) + 1):
            max_len = len(str(columns[col_idx - 1]))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    val_len = len(str(row_data[col_idx - 1]))
                    max_len = max(max_len, val_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(filepath)
    return filepath


def _append_correction_sheet(
    excel_path: str,
    correction_factors: dict,
    summary_data: list[tuple],
) -> str:
    ls = correction_factors.get("LS", 0.0)
    is_val = correction_factors.get("IS", 0.0)
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = "보정계수반영 대손율"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="맑은 고딕", size=10)
    title_font = Font(name="맑은 고딕", bold=True, size=14, color="1F3864")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"보정계수반영 대손율 (LS={ls}, IS={is_val})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    columns = ["구분", "대손비용률(%)", "보정계수", "LS", "IS", "LS*보정계수", "IS*보정계수", "보정계수반영 대손율(%)"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, (grp, cost_rate, corr) in enumerate(summary_data, 4):
        ls_corr = round(ls * corr, 4) if corr else 0
        is_corr = round(is_val * corr, 4) if corr else 0
        adj_rate = round(ls_corr + is_corr, 4)
        row_values = [grp, cost_rate, corr, ls, is_val, ls_corr, is_corr, adj_rate]
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.font = Font(name="맑은 고딕", size=10)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.0000'
            else:
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left")
    for col_idx in range(1, len(columns) + 1):
        max_len = len(columns[col_idx - 1])
        for r in range(4, 4 + len(summary_data)):
            val_len = len(str(ws.cell(row=r, column=col_idx).value or ""))
            max_len = max(max_len, val_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 25)
    ws.freeze_panes = ws.cell(row=4, column=1)
    wb.save(excel_path)
    return excel_path


def _tool_fn_대손비용률(params: dict) -> dict:
    """대손비용률 분석 Tool의 fn. 4개 SQL 실행 + 엑셀 생성 + 답변까지 처리.
    다른 Tool의 fn은 SQL 문자열을 반환하지만, 이 Tool은 dict를 반환합니다.
    execute_tool 노드에서 isinstance(result, dict)로 분기합니다."""
    merchant = params["가맹점명"]
    yyyymm = params["기준년월"]
    ls = float(params["LS"])
    is_val = float(params["IS"])

    query_results: dict[str, tuple[list[str], list[tuple]]] = {}
    all_sqls = []
    all_errors = []

    for sheet_name, template in BAD_DEBT_QUERIES.items():
        sql = _build_bad_debt_sql(template, params)
        all_sqls.append(f"-- [{sheet_name}]\n{sql}")
        columns, rows, error = execute_sql(sql)
        if error:
            all_errors.append(f"[{sheet_name}] {error}")
            query_results[sheet_name] = ([], [])
        else:
            query_results[sheet_name] = (columns, rows)

    combined_sql = "\n\n".join(all_sqls)

    if all(len(r[1]) == 0 for r in query_results.values()):
        return {
            "is_complete": True,
            "sql": combined_sql,
            "columns": [],
            "rows": [],
            "answer": (
                f"'{merchant}' 기업의 {yyyymm} 기준 대손비용률 데이터가 조회되지 않았습니다.\n"
                f"기업(가맹점)명이나 기간을 확인해주세요."
                + (f"\n\n오류 정보:\n" + "\n".join(all_errors) if all_errors else "")
            ),
            "excel_path": "",
        }

    excel_path = _generate_bad_debt_excel(params, query_results)

    summary_cols, summary_rows = query_results.get("대손비용률종합", ([], []))
    summary_for_correction = []
    if summary_cols and summary_rows:
        col_map = {c: i for i, c in enumerate(summary_cols)}
        for row in summary_rows:
            grp = row[col_map.get("구분", 0)] if "구분" in col_map else ""
            # 보정계수/대손비용률은 구분·잔액 조건에 따라 SQL에서 NULL이 나올 수 있다.
            cost_rate = _to_float(row[col_map["대손비용률_퍼센트"]]) if "대손비용률_퍼센트" in col_map else 0.0
            corr = _to_float(row[col_map["보정계수"]]) if "보정계수" in col_map else 0.0
            summary_for_correction.append((grp, cost_rate, corr))

    if summary_for_correction:
        correction = {"LS": ls, "IS": is_val}
        _append_correction_sheet(excel_path, correction, summary_for_correction)

    prev_month = _calc_months_back(yyyymm, 1)
    summary_parts = []
    for sheet_name in ["월초충당금", "월말충당금", "상각내역", "대손비용률종합"]:
        columns, rows = query_results.get(sheet_name, ([], []))
        if columns and rows:
            header = " | ".join(columns)
            data_lines = [" | ".join(str(v) for v in row) for row in rows[:20]]
            summary_parts.append(f"### {sheet_name}\n{header}\n" + "\n".join(data_lines))

    corr_summary = ""
    if summary_for_correction:
        corr_lines = []
        for grp, cost_rate, corr in summary_for_correction:
            ls_c = round(ls * corr, 4)
            is_c = round(is_val * corr, 4)
            adj = round(ls_c + is_c, 4)
            corr_lines.append(f"구분={grp}: LS*보정계수={ls_c}, IS*보정계수={is_c}, 보정계수반영 대손율={adj}%")
        corr_summary = "\n### 보정계수반영 대손율\n" + "\n".join(corr_lines)

    combined_data = "\n\n".join(summary_parts) + corr_summary if summary_parts else "데이터 없음"

    answer_prompt = f"""당신은 KB카드 기업영업 데이터 분석가입니다.
아래 쿼리 결과를 종합하여 '{merchant}'의 {yyyymm} 대손비용률 분석 결과를 짧고 직관적으로 작성하세요.

## 대손비용률 계산 로직
- 월초 충당금: 전월말({prev_month}) 기대대손충당금 및 원화대출잔액
- 월말 충당금: 당월말({yyyymm}) 기대대손충당금 및 원화대출잔액
- 상각금액: 해당 기간 특수채권편입원금 + 특수채권편입가지급금
- 대손비용 = 월말 충당금 - 월초 충당금 + 상각금액
- 대손비용률(%) = 대손비용 * 12 / Average(월초 잔액, 월말 잔액) * 100
- 전체 대손비용률(%) = SUM(구분0 + 구분1 대손비용) * 12 / Average(SUM(월초 잔액), SUM(월말 잔액)) * 100
- 보정계수 = 구분1 대손비용률 / 전체 대손비용률
- 보정계수반영 대손율 = LS * 보정계수 + IS * 보정계수  (LS={ls}, IS={is_val})

## 쿼리 결과
{combined_data}

## 답변 형식
아래 형식을 반드시 지키세요. 전체 답변은 12줄 이내로 제한하세요.

### 핵심 요약
| 항목 | 값 |
|---|---:|
| 해당 가맹점 구분='1' 대손비용률 | ... |
| 보정계수 | ... |
| 보정계수반영 대손율 | ... |

### 판단
- 가장 중요한 해석 1~2개만 작성하세요.

### 파일
- 상세 산출 내역은 엑셀 파일을 확인하라고 한 줄로 안내하세요.

## 작성 규칙
1. 금액은 억원/만원 단위로 변환하세요.
2. 비율은 소수점 4자리까지 %로 표시하세요.
3. 장황한 계산 과정 설명은 생략하고 표와 핵심 해석만 남기세요.
4. 구분='1'은 해당 가맹점 관련 고객, '0'은 비관련 고객입니다.
5. 데이터에 없는 내용은 추정하지 마세요.

답변:"""

    try:
        answer = _normalize_llm_text(_call_llm(answer_prompt))
    except Exception:
        answer = f"'{merchant}' {yyyymm} 대손비용률 분석 결과가 엑셀 파일로 생성되었습니다."
    answer += f"\n\n상세 분석 엑셀 파일: {excel_path}"

    return {
        "is_complete": True,
        "sql": combined_sql,
        "columns": summary_cols,
        "rows": summary_rows,
        "answer": answer,
        "excel_path": excel_path,
    }
